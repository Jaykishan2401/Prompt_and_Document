from docx import Document
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook


# PDF files will be processed using PyPDF2. Only first page of the pdf will be processed in accordance to word limit.
def pdfTextExtractor(path):
    reader = PdfReader(path)
    page = reader.pages[0]
    return page.extract_text()


# XLSX files will be processed using openxyl. I will be processing the [0][0] cell
def xlxsExtractor(path):
    wb = load_workbook(filename=path)
    sheet_obj = wb.active
    cell_obj = sheet_obj.cell(row=1, column=1)
    return cell_obj.value


# CSV files are processed using pandas. Will only return the head of the file.
def csvExtractor(path):
    df = pd.read_csv(path)
    text = df.head()
    return text

# DOCX files will be processed using python-docx module. Will only consider the first paragraph for brevity.
def wordExtractor(path):
    doc = Document(path)
    text = doc.paragraphs[0].text
    return text
